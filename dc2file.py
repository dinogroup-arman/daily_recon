import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook


# Read the Excel file without setting headers
file_path = 'mirae_account_position_summary.xlsx'
df = pd.read_excel(file_path, header=None, engine='openpyxl')

# Drop the first 3 rows
df = df.iloc[3:]

# Set the next row (previously the 4th row) as headers
df.columns = df.iloc[0]
df = df.drop(df.index[0])

# Define columns to drop and drop them
cols_to_drop = ['Cusip', 'Currency', 'Description', 'Coupon', 'Maturity Date', 'Open Position',
                'Open Avg Price', 'Open Position Interest Per Unit', 'Close Position',
                'Close Avg Price', 'Close Position Interest Per Unit', 'Market Price',
                'Pool Factor', 'No Of Tickets']

df = df.drop(columns=cols_to_drop)

# Define columns to sum up
cols_to_sum = ['PnL', 'Bond Interest', 'Principle Pay Down', 'Premium', 'Bond Coupon', 'UnRealized PnL']

# Group by 'Account' and sum the specified columns
result = df.groupby('Account')[cols_to_sum].sum().reset_index()

# Save the result to a new Excel file
result_file_path = 'summed_values_by_account.xlsx'
result.to_excel(result_file_path, index=False, engine='openpyxl')

# Define a style for the desired number format
number_format_style = NamedStyle(name="number_format_style", number_format="#,##0.00_);[Red](#,##0.00)")

# Apply the number_format_style to the specified columns in the Excel file
wb = Workbook()
wb = load_workbook(result_file_path)
ws = wb.active

# Get the number of rows and columns in the worksheet
max_row = ws.max_row
max_col = ws.max_column

# Iterate over the columns and rows and apply the style
for col in range(2, max_col + 1):  # Starting from column 2 (column B) as column A is 'Account'
    for row in range(2, max_row + 1):  # Starting from row 2 as the first row is the header
        ws.cell(row=row, column=col).style = number_format_style

wb.save(result_file_path)

print("Summed values by account with style applied saved to:", result_file_path)
